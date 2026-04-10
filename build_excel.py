"""Generate the Excel market-survey workbook.

Reads ``market_data.json`` (produced by ``scrape_market.py``) and writes
``market_survey_report.xlsx``. The workbook contains:

  - A ``Summary`` sheet with metadata and per-bedroom-type summary tables
  - One sheet per bedroom type (Studios, One-Bedrooms, Two-Bedrooms,
    Three-Bedrooms) containing a scatter chart followed by a formatted
    data table

Uses openpyxl's ScatterChart for all charts and the shared
``PROPERTY_COLORS`` palette from ``config.py`` for consistency with the
HTML dashboard.
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    BEDROOM_SECTIONS,
    LIGHT_GRAY_HEX,
    MARKET_PROPERTIES,
    NAVY_HEX,
    PROPERTY_COLORS,
    SUBJECT_PROPERTY,
    WHITE_HEX,
)

INPUT_JSON = Path("market_data.json")
OUTPUT_XLSX = Path("market_survey_report.xlsx")


# ---------------------------------------------------------------------------
# Styling primitives
# ---------------------------------------------------------------------------


NAVY_FILL = PatternFill("solid", fgColor=NAVY_HEX)
GRAY_FILL = PatternFill("solid", fgColor=LIGHT_GRAY_HEX)
WHITE_FILL = PatternFill("solid", fgColor=WHITE_HEX)
WHITE_BOLD = Font(color=WHITE_HEX, bold=True, name="Calibri")
WHITE_BOLD_LG = Font(color=WHITE_HEX, bold=True, size=14, name="Calibri")
NAVY_BOLD = Font(color=NAVY_HEX, bold=True, name="Calibri")
NAVY_LINK = Font(color=NAVY_HEX, underline="single", name="Calibri")
GRAY_ITALIC = Font(color="808080", italic=True, name="Calibri")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def hex_no_hash(h: str) -> str:
    return h.lstrip("#").upper()


def left_accent_border(color_hex: str) -> Border:
    return Border(left=Side(style="thick", color=hex_no_hash(color_hex)))


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------


def load_records() -> list[dict[str, Any]]:
    if not INPUT_JSON.exists():
        print(f"[error] {INPUT_JSON} not found. Run scrape_market.py first.")
        sys.exit(1)
    return json.loads(INPUT_JSON.read_text())


def live_units(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [r for r in records if not r.get("no_data")]


def units_for_bedroom(
    records: list[dict[str, Any]], bedrooms: int
) -> list[dict[str, Any]]:
    return [r for r in live_units(records) if r.get("bedrooms") == bedrooms]


def sort_property_order(names: list[str]) -> list[str]:
    rest = sorted(n for n in names if n != SUBJECT_PROPERTY)
    if SUBJECT_PROPERTY in names:
        return [SUBJECT_PROPERTY, *rest]
    return rest


def property_url_map() -> dict[str, str]:
    return {p["name"]: p["url"] for p in MARKET_PROPERTIES}


# ---------------------------------------------------------------------------
# Scatter chart builder
# ---------------------------------------------------------------------------


def _style_series(series: Series, color_hex: str, is_subject: bool) -> None:
    color = hex_no_hash(color_hex)
    line_width = 28575 if is_subject else 19050  # thicker line for subject
    line_props = LineProperties(
        w=line_width,
        solidFill=ColorChoice(srgbClr=color),
    )
    series.graphicalProperties = GraphicalProperties(ln=line_props)
    series.graphicalProperties.solidFill = color
    marker = Marker(symbol="circle", size=9 if is_subject else 7)
    marker.graphicalProperties = GraphicalProperties(
        solidFill=color,
        ln=LineProperties(solidFill=ColorChoice(srgbClr=color)),
    )
    series.marker = marker


def build_scatter_chart(
    ws: Worksheet,
    section_units: list[dict[str, Any]],
    section_title: str,
) -> ScatterChart | None:
    """Build a ScatterChart for a bedroom-type sheet.

    Data is written to a helper area on the sheet (columns far to the
    right, starting at column K) so openpyxl has valid Reference ranges
    to build each series from.
    """
    if not section_units:
        return None

    by_property: dict[str, list[dict[str, Any]]] = {}
    for u in section_units:
        by_property.setdefault(u["property_name"], []).append(u)

    ordered = sort_property_order(list(by_property.keys()))

    chart = ScatterChart()
    chart.title = section_title
    chart.style = 2
    chart.x_axis.title = "Square Feet"
    chart.y_axis.title = "Rental Rates"
    chart.y_axis.number_format = "$#,##0"
    chart.legend.position = "b"
    chart.width = 25
    chart.height = 15

    # Write the helper data block far to the right to stay out of the
    # visible table area. Column layout per property: [SqFt, Rent]
    helper_start_col = 11  # K
    current_col = helper_start_col

    for name in ordered:
        points = sorted(
            by_property[name], key=lambda u: u.get("sqft") or 0
        )
        points = [p for p in points if p.get("sqft") and p.get("asking_rent")]
        if not points:
            continue

        x_col = current_col
        y_col = current_col + 1
        # Header
        ws.cell(row=1, column=x_col, value=f"{name} SF")
        ws.cell(row=1, column=y_col, value=f"{name} Rent")
        for i, p in enumerate(points, start=2):
            ws.cell(row=i, column=x_col, value=p["sqft"])
            ws.cell(row=i, column=y_col, value=p["asking_rent"])

        x_ref = Reference(
            ws, min_col=x_col, min_row=2, max_col=x_col, max_row=1 + len(points)
        )
        y_ref = Reference(
            ws, min_col=y_col, min_row=2, max_col=y_col, max_row=1 + len(points)
        )
        series = Series(y_ref, xvalues=x_ref, title=name)
        _style_series(
            series,
            PROPERTY_COLORS.get(name, "#777777"),
            is_subject=(name == SUBJECT_PROPERTY),
        )
        chart.series.append(series)
        current_col += 2

    # Hide the helper columns so the workbook looks clean
    for col_idx in range(helper_start_col, current_col):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].hidden = True

    return chart


# ---------------------------------------------------------------------------
# Bedroom-type sheet
# ---------------------------------------------------------------------------


DATA_HEADERS = [
    "Property Name",
    "Unit",
    "Date Available",
    "Concession",
    "Floorplan",
    "Sq Ft",
    "Asking Rent",
    "PSF",
]


def build_bedroom_sheet(
    wb: Workbook,
    sheet_name: str,
    section_title: str,
    section_label_title: str,
    section_units: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(sheet_name)
    prop_url = property_url_map()

    # Row 1: title bar merged A:H
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{sheet_name.upper()} — Market Survey Comparison"
    title_cell.fill = NAVY_FILL
    title_cell.font = WHITE_BOLD_LG
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Column widths per spec
    widths = {"A": 28, "B": 10, "C": 16, "D": 16, "E": 14, "F": 10, "G": 14, "H": 10}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    # Scatter chart anchored at A3, ~20 rows tall
    chart = build_scatter_chart(ws, section_units, section_title)
    if chart is not None:
        ws.add_chart(chart, "A3")

    # Data table starts ~2 rows below the 20-row chart (row 3 + 20 = 23; +2 = 25)
    data_start_row = 25

    # Table title row merged A:H
    title_row = data_start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)
    t = ws.cell(row=title_row, column=1)
    t.value = section_label_title.upper()
    t.fill = NAVY_FILL
    t.font = WHITE_BOLD
    t.alignment = CENTER
    ws.row_dimensions[title_row].height = 22

    # Header row
    header_row = title_row + 1
    for col_idx, header in enumerate(DATA_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = NAVY_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Data rows — subject first, then alphabetical
    all_property_names = [p["name"] for p in MARKET_PROPERTIES]
    ordered = sort_property_order(all_property_names)
    by_property: dict[str, list[dict[str, Any]]] = {}
    for u in section_units:
        by_property.setdefault(u["property_name"], []).append(u)

    current_row = header_row + 1
    row_parity = 0

    for prop_name in ordered:
        color = PROPERTY_COLORS.get(prop_name, "#777777")
        url = prop_url.get(prop_name, "")
        is_subject = prop_name == SUBJECT_PROPERTY
        prop_units = sorted(
            by_property.get(prop_name, []),
            key=lambda u: u.get("sqft") or 0,
        )

        if not prop_units:
            # Single "No data available" row
            fill = WHITE_FILL if row_parity % 2 == 0 else GRAY_FILL
            a = ws.cell(row=current_row, column=1, value=prop_name)
            a.font = NAVY_BOLD if is_subject else Font(name="Calibri")
            a.alignment = LEFT
            a.fill = fill
            a.border = left_accent_border(color)
            # Merge B:H and write "No data available"
            ws.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=8,
            )
            no_data = ws.cell(row=current_row, column=2, value="No data available")
            no_data.font = GRAY_ITALIC
            no_data.alignment = CENTER
            no_data.fill = fill
            # Fill adjacent cells to keep the merged fill consistent
            for col_idx in range(3, 9):
                ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
            row_parity += 1
            continue

        for u in prop_units:
            fill = WHITE_FILL if row_parity % 2 == 0 else GRAY_FILL

            # Column A — hyperlink via HYPERLINK() formula
            a = ws.cell(row=current_row, column=1)
            if url:
                a.value = f'=HYPERLINK("{url}","{prop_name}")'
            else:
                a.value = prop_name
            a.font = (
                Font(
                    color=NAVY_HEX,
                    underline="single",
                    bold=True,
                    name="Calibri",
                )
                if is_subject
                else NAVY_LINK
            )
            a.alignment = LEFT
            a.fill = fill
            a.border = left_accent_border(color)

            # B: Unit
            b = ws.cell(row=current_row, column=2, value=u.get("unit_number") or "—")
            b.alignment = CENTER
            b.fill = fill

            # C: Date Available
            c = ws.cell(row=current_row, column=3, value=u.get("date_available") or "—")
            c.alignment = CENTER
            c.fill = fill

            # D: Concession
            d = ws.cell(
                row=current_row,
                column=4,
                value=u.get("concession") if u.get("concession") else "None",
            )
            d.alignment = CENTER
            d.fill = fill

            # E: Floorplan
            e = ws.cell(row=current_row, column=5, value=u.get("floorplan_name") or "—")
            e.alignment = CENTER
            e.fill = fill

            # F: Sq Ft
            f = ws.cell(row=current_row, column=6, value=u.get("sqft"))
            f.number_format = "#,##0"
            f.alignment = RIGHT
            f.fill = fill

            # G: Asking Rent
            g = ws.cell(row=current_row, column=7, value=u.get("asking_rent"))
            g.number_format = "$#,##0"
            g.alignment = RIGHT
            g.fill = fill

            # H: PSF
            h = ws.cell(row=current_row, column=8, value=u.get("psf"))
            h.number_format = "$#,##0.00"
            h.alignment = RIGHT
            h.fill = fill

            current_row += 1
            row_parity += 1


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------


def _set_row(
    ws: Worksheet,
    row: int,
    values: list[Any],
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    align: Alignment | None = None,
    number_formats: dict[int, str] | None = None,
) -> None:
    number_formats = number_formats or {}
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if col_idx in number_formats:
            cell.number_format = number_formats[col_idx]


def build_summary_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Summary", 0)

    widths = {"A": 28, "B": 18, "C": 14, "D": 14, "E": 14, "F": 12, "G": 12}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    # Row 1: title merged A:G
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Market Survey — McKinney / Allen"
    c.fill = NAVY_FILL
    c.font = WHITE_BOLD_LG
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    live = live_units(records)
    prop_with_data = {r["property_name"] for r in live}
    total_tracked = len(live)

    today = date.today().strftime("%B %d, %Y")
    metadata = [
        ("Scraped:", today),
        ("Subject Property:", SUBJECT_PROPERTY),
        ("Submarket:", "McKinney / Allen, DFW"),
        ("Total Units Tracked:", total_tracked),
        ("Properties with Data:", f"{len(prop_with_data)} of {len(MARKET_PROPERTIES)}"),
    ]
    bold = Font(bold=True, name="Calibri")
    for i, (label, value) in enumerate(metadata, start=3):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value)

    # Per-bedroom-type summary tables stacked below
    current_row = 3 + len(metadata) + 2  # gap row after metadata

    for cfg in BEDROOM_SECTIONS:
        section_units = units_for_bedroom(records, cfg["bedrooms"])

        # Title row merged A:G
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=7
        )
        t = ws.cell(row=current_row, column=1, value=cfg["label"])
        t.fill = NAVY_FILL
        t.font = WHITE_BOLD
        t.alignment = CENTER
        current_row += 1

        # Header row
        headers = [
            "Property Name",
            "# Units Available",
            "Min Rent",
            "Max Rent",
            "Avg Rent",
            "Avg PSF",
            "Avg SF",
        ]
        _set_row(
            ws,
            current_row,
            headers,
            fill=NAVY_FILL,
            font=WHITE_BOLD,
            align=CENTER,
        )
        current_row += 1

        # One row per property (all 7)
        all_rents: list[float] = []
        all_psf: list[float] = []
        all_sf: list[float] = []
        total_units = 0

        ordered = sort_property_order([p["name"] for p in MARKET_PROPERTIES])
        for prop_name in ordered:
            is_subject = prop_name == SUBJECT_PROPERTY
            prop_units = [u for u in section_units if u["property_name"] == prop_name]
            row_font = NAVY_BOLD if is_subject else Font(name="Calibri")

            if not prop_units:
                values: list[Any] = [prop_name, 0, "—", "—", "—", "—", "—"]
                _set_row(ws, current_row, values, font=row_font)
            else:
                rents = [u["asking_rent"] for u in prop_units if u.get("asking_rent")]
                psfs = [u["psf"] for u in prop_units if u.get("psf")]
                sfs = [u["sqft"] for u in prop_units if u.get("sqft")]
                values = [
                    prop_name,
                    len(prop_units),
                    min(rents) if rents else "—",
                    max(rents) if rents else "—",
                    round(mean(rents)) if rents else "—",
                    round(mean(psfs), 2) if psfs else "—",
                    round(mean(sfs)) if sfs else "—",
                ]
                _set_row(
                    ws,
                    current_row,
                    values,
                    font=row_font,
                    number_formats={
                        3: "$#,##0",
                        4: "$#,##0",
                        5: "$#,##0",
                        6: "$#,##0.00",
                        7: "#,##0",
                    },
                )
                total_units += len(prop_units)
                all_rents.extend(rents)
                all_psf.extend(psfs)
                all_sf.extend(sfs)
            current_row += 1

        # Totals row
        totals_values: list[Any] = [
            "Totals / Averages",
            total_units,
            min(all_rents) if all_rents else "—",
            max(all_rents) if all_rents else "—",
            round(mean(all_rents)) if all_rents else "—",
            round(mean(all_psf), 2) if all_psf else "—",
            round(mean(all_sf)) if all_sf else "—",
        ]
        _set_row(
            ws,
            current_row,
            totals_values,
            fill=GRAY_FILL,
            font=bold,
            number_formats={
                3: "$#,##0",
                4: "$#,##0",
                5: "$#,##0",
                6: "$#,##0.00",
                7: "#,##0",
            },
        )
        current_row += 2  # gap before next bedroom-type summary table


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    records = load_records()
    wb = Workbook()
    # Remove the default blank sheet openpyxl creates
    default = wb.active
    wb.remove(default)

    # Build each bedroom-type sheet (skip empty ones entirely)
    for cfg in BEDROOM_SECTIONS:
        section_units = units_for_bedroom(records, cfg["bedrooms"])
        if not section_units:
            continue
        build_bedroom_sheet(
            wb,
            sheet_name=cfg["label"],
            section_title=cfg["section_title"],
            section_label_title=cfg["label"],
            section_units=section_units,
        )

    # Summary sheet (inserted as the first tab)
    build_summary_sheet(wb, records)

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

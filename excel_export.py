"""Excel export module — generates a workbook matching the reference format.

Creates 5 tabs:
  Summary  – property-level stats by unit type
  Studios  – scatter chart + unit list
  1BR      – scatter chart + unit list
  2BR      – scatter chart + unit list
  3BR      – scatter chart + unit list
"""

import io
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import UNIT_TYPES, HEADER_FILL_COLOR, HEADER_FONT_COLOR


# -- Consistent property color palette for charts --
PROPERTY_COLORS = [
    "C00000",  # dark red
    "0070C0",  # blue
    "00B050",  # green
    "7030A0",  # purple
    "FF6600",  # orange
    "00B0F0",  # light blue
    "FFC000",  # gold
]

HEADER_FILL = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=HEADER_FONT_COLOR)
DATA_FONT = Font(name="Calibri", size=10)
MONEY_FORMAT = '$#,##0'
PSF_FORMAT = '$#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True)


def build_workbook(df: pd.DataFrame) -> Workbook:
    """Build the full Excel workbook from a DataFrame of units."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Build tabs
    _build_summary_tab(wb, df)
    for unit_type in UNIT_TYPES:
        _build_unit_type_tab(wb, df, unit_type)

    return wb


def save_workbook(df: pd.DataFrame, filepath: str) -> str:
    """Build and save the workbook to disk."""
    wb = build_workbook(df)
    wb.save(filepath)
    return filepath


def workbook_to_bytes(df: pd.DataFrame) -> bytes:
    """Build workbook and return as bytes (for Streamlit download)."""
    wb = build_workbook(df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Summary Tab
# ---------------------------------------------------------------------------

def _build_summary_tab(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Summary")

    # --- TITLE ROW ---
    ws.merge_cells("A2:I2")
    ws["A2"] = "SUMMARY"
    ws["A2"].font = TITLE_FONT
    ws["A2"].fill = TITLE_FILL
    ws["A2"].alignment = Alignment(horizontal="center")
    for col in range(1, 10):
        ws.cell(row=2, column=col).fill = TITLE_FILL

    ws.merge_cells("K2:W2")
    date_str = datetime.now().strftime("%m/%d/%Y")
    ws["K2"] = f"AVERAGE MARKET RENTS  AS OF {date_str}"
    ws["K2"].font = TITLE_FONT
    ws["K2"].fill = TITLE_FILL
    ws["K2"].alignment = Alignment(horizontal="center")
    for col in range(11, 24):
        ws.cell(row=2, column=col).fill = TITLE_FILL

    # --- HEADER ROW ---
    summary_headers = ["PROPERTY", "Studio", "1BR", "2BR", "3BR", "TOTAL", "AVG SQ FT", "", "",
                        "Studio MKT", "Studio PSF", "", "1BR MKT", "1BR PSF", "",
                        "2BR MKT", "2BR PSF", "", "3BR MKT", "3BR PSF"]
    for i, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # --- PROPERTY ROWS ---
    properties = df["Property Name"].unique()
    row = 4

    for prop in properties:
        pdf = df[df["Property Name"] == prop]
        ws.cell(row=row, column=1, value=prop).font = Font(name="Calibri", size=10, color="0070C0")

        for col_idx, utype in enumerate(UNIT_TYPES, 2):
            count = len(pdf[pdf["Unit Type"] == utype])
            ws.cell(row=row, column=col_idx, value=count if count > 0 else 0).font = DATA_FONT

        total = len(pdf)
        ws.cell(row=row, column=6, value=total).font = DATA_FONT

        avg_sqft = int(pdf["Sq Ft"].mean()) if len(pdf) > 0 and pdf["Sq Ft"].mean() > 0 else 0
        ws.cell(row=row, column=7, value=avg_sqft).font = DATA_FONT

        # Market rents by unit type
        rent_col_map = {
            "Studios": (10, 11),
            "1BR": (13, 14),
            "2BR": (16, 17),
            "3BR": (19, 20),
        }
        for utype, (mkt_col, psf_col) in rent_col_map.items():
            type_df = pdf[pdf["Unit Type"] == utype]
            if len(type_df) > 0:
                avg_rent = type_df["Asking Rent"].mean()
                avg_psf = type_df["PSF"].mean() if "PSF" in type_df.columns else 0
                if pd.notna(avg_rent) and avg_rent > 0:
                    cell_mkt = ws.cell(row=row, column=mkt_col, value=round(avg_rent))
                    cell_mkt.number_format = MONEY_FORMAT
                    cell_mkt.font = DATA_FONT
                if pd.notna(avg_psf) and avg_psf > 0:
                    cell_psf = ws.cell(row=row, column=psf_col, value=round(avg_psf, 2))
                    cell_psf.number_format = PSF_FORMAT
                    cell_psf.font = DATA_FONT

        # Apply borders
        for col in range(1, 21):
            ws.cell(row=row, column=col).border = THIN_BORDER

        row += 1

    # --- WEIGHTED AVERAGE / TOTAL ROW ---
    row += 1
    ws.cell(row=row, column=1, value="Wtd Avg / Total").font = SUBTOTAL_FONT

    for col_idx, utype in enumerate(UNIT_TYPES, 2):
        count = len(df[df["Unit Type"] == utype])
        cell = ws.cell(row=row, column=col_idx, value=count)
        cell.font = SUBTOTAL_FONT
        cell.fill = SUBTOTAL_FILL

    total = len(df)
    ws.cell(row=row, column=6, value=total).font = SUBTOTAL_FONT
    ws.cell(row=row, column=6).fill = SUBTOTAL_FILL

    avg_sqft = int(df["Sq Ft"].mean()) if len(df) > 0 else 0
    ws.cell(row=row, column=7, value=avg_sqft).font = SUBTOTAL_FONT
    ws.cell(row=row, column=7).fill = SUBTOTAL_FILL

    for utype, (mkt_col, psf_col) in rent_col_map.items():
        type_df = df[df["Unit Type"] == utype]
        if len(type_df) > 0:
            avg_rent = type_df["Asking Rent"].mean()
            avg_psf = type_df["PSF"].mean()
            if pd.notna(avg_rent) and avg_rent > 0:
                cell = ws.cell(row=row, column=mkt_col, value=round(avg_rent))
                cell.number_format = MONEY_FORMAT
                cell.font = SUBTOTAL_FONT
                cell.fill = SUBTOTAL_FILL
            if pd.notna(avg_psf) and avg_psf > 0:
                cell = ws.cell(row=row, column=psf_col, value=round(avg_psf, 2))
                cell.number_format = PSF_FORMAT
                cell.font = SUBTOTAL_FONT
                cell.fill = SUBTOTAL_FILL

    for col in range(1, 21):
        ws.cell(row=row, column=col).border = THIN_BORDER

    # --- McKinney Terrace standalone row ---
    row += 2
    mkt_df = df[df["Property Name"] == "McKinney Terrace"]
    if len(mkt_df) > 0:
        ws.cell(row=row, column=1, value="McKinney Terrace").font = Font(name="Calibri", size=10, bold=True)
        for col_idx, utype in enumerate(UNIT_TYPES, 2):
            count = len(mkt_df[mkt_df["Unit Type"] == utype])
            ws.cell(row=row, column=col_idx, value=count).font = DATA_FONT

        total = len(mkt_df)
        ws.cell(row=row, column=6, value=total).font = DATA_FONT
        avg_sqft = int(mkt_df["Sq Ft"].mean()) if len(mkt_df) > 0 else 0
        ws.cell(row=row, column=7, value=avg_sqft).font = DATA_FONT

        for utype, (mkt_col, psf_col) in rent_col_map.items():
            type_df = mkt_df[mkt_df["Unit Type"] == utype]
            if len(type_df) > 0:
                avg_rent = type_df["Asking Rent"].mean()
                avg_psf = type_df["PSF"].mean()
                if pd.notna(avg_rent) and avg_rent > 0:
                    cell = ws.cell(row=row, column=mkt_col, value=round(avg_rent))
                    cell.number_format = MONEY_FORMAT
                    cell.font = DATA_FONT
                if pd.notna(avg_psf) and avg_psf > 0:
                    cell = ws.cell(row=row, column=psf_col, value=round(avg_psf, 2))
                    cell.number_format = PSF_FORMAT
                    cell.font = DATA_FONT

        for col in range(1, 21):
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 22
    for c in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 10
    for c in range(10, 21):
        ws.column_dimensions[get_column_letter(c)].width = 12


# ---------------------------------------------------------------------------
# Unit-Type Tabs (Studios / 1BR / 2BR / 3BR)
# ---------------------------------------------------------------------------

def _build_unit_type_tab(wb: Workbook, df: pd.DataFrame, unit_type: str):
    ws = wb.create_sheet(unit_type)
    type_df = df[df["Unit Type"] == unit_type].copy()
    type_df = type_df.sort_values(["Property Name", "Sq Ft"]).reset_index(drop=True)

    properties = type_df["Property Name"].unique().tolist()

    # --- TITLE ---
    ws.merge_cells("A2:I2")
    beds_label = {"Studios": "Studios", "1BR": "1-Bedrooms", "2BR": "2-Bedrooms", "3BR": "3-Bedrooms"}
    ws["A2"] = beds_label.get(unit_type, unit_type)
    ws["A2"].font = TITLE_FONT
    ws["A2"].fill = TITLE_FILL
    ws["A2"].alignment = Alignment(horizontal="left")
    for col in range(1, 10):
        ws.cell(row=2, column=col).fill = TITLE_FILL

    ws.merge_cells("A3:I3")
    ws["A3"] = "Market Survey Comparison"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
    ws["A3"].fill = TITLE_FILL
    for col in range(1, 10):
        ws.cell(row=3, column=col).fill = TITLE_FILL

    # --- SCATTER CHART ---
    chart = ScatterChart()
    chart.title = None
    chart.x_axis.title = "Square Feet"
    chart.y_axis.title = "Rental Rates"
    chart.x_axis.numFmt = '#,##0'
    chart.y_axis.numFmt = '$#,##0'
    chart.style = 2
    chart.width = 28
    chart.height = 16
    chart.legend.position = "b"

    # We need to write chart data to hidden columns or use inline data.
    # Write per-property data starting at column K for chart references.
    chart_data_start_col = 11  # Column K
    chart_row_start = 4

    for prop_idx, prop in enumerate(properties):
        prop_df = type_df[type_df["Property Name"] == prop].copy()
        if len(prop_df) == 0:
            continue

        # Write sqft and rent for this property
        col_sqft = chart_data_start_col + (prop_idx * 2)
        col_rent = col_sqft + 1

        # Header
        ws.cell(row=chart_row_start, column=col_sqft, value=f"{prop} SqFt")
        ws.cell(row=chart_row_start, column=col_rent, value=f"{prop} Rent")

        for i, (_, unit_row) in enumerate(prop_df.iterrows()):
            data_row = chart_row_start + 1 + i
            sqft_val = unit_row.get("Sq Ft", 0)
            rent_val = unit_row.get("Asking Rent", 0)
            if pd.notna(sqft_val) and pd.notna(rent_val) and sqft_val > 0 and rent_val > 0:
                ws.cell(row=data_row, column=col_sqft, value=sqft_val)
                ws.cell(row=data_row, column=col_rent, value=rent_val)

        # Create series
        max_data_row = chart_row_start + len(prop_df)
        x_values = Reference(ws, min_col=col_sqft, min_row=chart_row_start + 1,
                             max_row=max_data_row)
        y_values = Reference(ws, min_col=col_rent, min_row=chart_row_start + 1,
                             max_row=max_data_row)

        series = Series(y_values, x_values, title=prop)
        series.graphicalProperties.line.noFill = True  # scatter, no lines

        # Set marker color
        color_hex = PROPERTY_COLORS[prop_idx % len(PROPERTY_COLORS)]
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = color_hex

        chart.series.append(series)

    # Add trendline (linear) for overall market using the first series approach
    # The chart itself shows the scatter of all properties

    ws.add_chart(chart, "A4")

    # --- UNIT TYPE LABEL ---
    table_title_row = 36
    ws.merge_cells(f"A{table_title_row}:I{table_title_row}")
    ws.cell(row=table_title_row, column=1, value=unit_type.upper())
    ws.cell(row=table_title_row, column=1).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=table_title_row, column=1).alignment = Alignment(horizontal="center")

    # --- TABLE HEADERS ---
    header_row = table_title_row + 1
    table_headers = [
        "Property Name", "Unit", "Date Available", "Concession",
        "Floorplan", "Sq Ft", "0/2026 |Gross Re", "PSF",
    ]
    col_widths = [22, 10, 16, 14, 12, 8, 16, 8]

    for i, (h, w) in enumerate(zip(table_headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- DATA ROWS ---
    data_row = header_row + 1
    for _, unit_row in type_df.iterrows():
        prop_name = unit_row.get("Property Name", "")
        unit_num = unit_row.get("Unit", "")
        avail = unit_row.get("Move-In Date", "")
        concession = unit_row.get("Concessions", "")
        floorplan = unit_row.get("Floorplan", "")
        sqft = unit_row.get("Sq Ft", 0)
        rent = unit_row.get("Asking Rent", 0)
        psf = unit_row.get("PSF", 0)

        # Property name with hyperlink color
        cell_prop = ws.cell(row=data_row, column=1, value=prop_name)
        cell_prop.font = Font(name="Calibri", size=10, color="0070C0")
        cell_prop.border = THIN_BORDER

        ws.cell(row=data_row, column=2, value=unit_num).font = DATA_FONT
        ws.cell(row=data_row, column=2).border = THIN_BORDER
        ws.cell(row=data_row, column=2).alignment = Alignment(horizontal="center")

        ws.cell(row=data_row, column=3, value=avail or "").font = DATA_FONT
        ws.cell(row=data_row, column=3).border = THIN_BORDER

        ws.cell(row=data_row, column=4, value=concession or "").font = DATA_FONT
        ws.cell(row=data_row, column=4).border = THIN_BORDER

        ws.cell(row=data_row, column=5, value=floorplan).font = DATA_FONT
        ws.cell(row=data_row, column=5).border = THIN_BORDER
        ws.cell(row=data_row, column=5).alignment = Alignment(horizontal="center")

        cell_sqft = ws.cell(row=data_row, column=6, value=sqft if pd.notna(sqft) else 0)
        cell_sqft.font = DATA_FONT
        cell_sqft.number_format = '#,##0'
        cell_sqft.border = THIN_BORDER
        cell_sqft.alignment = Alignment(horizontal="center")

        cell_rent = ws.cell(row=data_row, column=7, value=rent if pd.notna(rent) else 0)
        cell_rent.font = DATA_FONT
        cell_rent.number_format = MONEY_FORMAT
        cell_rent.border = THIN_BORDER

        cell_psf = ws.cell(row=data_row, column=8, value=psf if pd.notna(psf) else 0)
        cell_psf.font = DATA_FONT
        cell_psf.number_format = PSF_FORMAT
        cell_psf.border = THIN_BORDER

        data_row += 1

    # Hide chart data columns
    for col in range(chart_data_start_col, chart_data_start_col + len(properties) * 2 + 2):
        ws.column_dimensions[get_column_letter(col)].hidden = True
